import json
from io import BytesIO

import httpx
from demo_ingestion import access_token, seed_scope
from openpyxl import load_workbook
from pypdf import PdfReader


def main() -> None:
    subject, product_id, tenant_id, entity_id = seed_scope()
    headers = {
        "Authorization": f"Bearer {access_token(subject)}",
        "X-Product-ID": product_id,
        "X-Tenant-ID": tenant_id,
    }
    common = {
        "dataset": "attendance",
        "entity_id": entity_id,
        "module": "attendance",
        "classification": 1,
    }
    jobs = {}
    downloads: dict[str, bytes] = {}
    with httpx.Client(base_url="http://api:8000", timeout=60) as client:
        for format_name in ("json", "xlsx", "pdf"):
            response = client.post(
                "/api/v1/exports",
                headers=headers,
                json={**common, "format": format_name},
            )
            response.raise_for_status()
            jobs[format_name] = response.json()
            download = client.get(
                f"/api/v1/exports/{response.json()['export_id']}/download",
                headers=headers,
            )
            download.raise_for_status()
            downloads[format_name] = download.content

    json_rows = json.loads(downloads["json"])["records"]
    json_ids = {row["employee_id"] for row in json_rows}
    workbook = load_workbook(BytesIO(downloads["xlsx"]), read_only=True, data_only=False)
    sheet = workbook["Attendance"]
    header_row = [cell.value for cell in next(sheet.iter_rows())]
    employee_column = header_row.index("employee_id")
    xlsx_ids = {
        row[employee_column].value
        for row in sheet.iter_rows(min_row=2)
        if row[employee_column].value
    }
    pdf_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(BytesIO(downloads["pdf"])).pages
    )
    pdf_ids = {employee_id for employee_id in json_ids if employee_id in pdf_text}
    output = {
        "jobs": jobs,
        "byte_sizes": {name: len(content) for name, content in downloads.items()},
        "logical_employee_ids": sorted(json_ids),
        "consistent": json_ids == xlsx_ids == pdf_ids,
    }
    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
