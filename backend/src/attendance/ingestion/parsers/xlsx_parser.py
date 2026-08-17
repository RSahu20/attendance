from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from attendance.domain.attendance import ExtractionMethod
from attendance.ingestion.parsers.base import DocumentParser
from attendance.ingestion.parsers.tabular import AttendanceHeaderNotFound, table_units
from attendance.ingestion.types import IngestionError, ParsedUnit


class XLSXParser(DocumentParser):
    name = "xlsx"

    def parse(self, content: bytes, filename: str) -> list[ParsedUnit]:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            units: list[ParsedUnit] = []
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                if not any(any(value is not None for value in row) for row in rows):
                    continue
                try:
                    units.extend(
                        table_units(
                            rows,
                            filename=filename,
                            location={"sheet": sheet.title},
                            key_prefix=f"sheet:{sheet.title}",
                            method=ExtractionMethod.XLSX,
                        )
                    )
                except AttendanceHeaderNotFound:
                    continue
            workbook.close()
            if not units:
                raise IngestionError(
                    "no_attendance_header",
                    "The XLSX is readable, but no attendance header row was detected",
                )
            return units
        except IngestionError:
            raise
        except (InvalidFileException, OSError, ValueError, KeyError) as exc:
            raise IngestionError("invalid_xlsx", "The XLSX file could not be parsed") from exc
