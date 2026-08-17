import json
from collections.abc import Iterator
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

import jwt
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import Session

from attendance.api.routes.exports import get_export_storage
from attendance.config import get_settings
from attendance.db.models.identity import User
from attendance.db.models.rbac import UserRoleAssignment
from attendance.main import app
from attendance.providers.storage.local import LocalStorageProvider
from tests.integration.conftest import SecuritySeed


def _token(subject: str) -> str:
    settings = get_settings()
    now = datetime.now(UTC)
    return jwt.encode(
        {
            "sub": subject,
            "iat": now,
            "exp": now + timedelta(minutes=10),
            "iss": settings.jwt_issuer,
            "aud": settings.jwt_audience,
        },
        settings.jwt_secret.get_secret_value(),
        algorithm="HS256",
    )


def _headers(
    seed: SecuritySeed, *, subject: str | None = None, tenant: Any = None
) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {_token(subject or seed.subject)}",
        "X-Product-ID": str(seed.product_id),
        "X-Tenant-ID": str(tenant or seed.tenant_a_id),
    }


def _upload(
    client: TestClient,
    seed: SecuritySeed,
    content: bytes,
    *,
    logical_name: str,
    entity_id: Any,
    classification: int,
) -> None:
    response = client.post(
        "/api/v1/documents",
        headers=_headers(seed),
        files={"file": (f"{logical_name}.csv", content, "text/csv")},
        data={
            "entity_id": str(entity_id),
            "module": "attendance",
            "classification": str(classification),
            "logical_name": logical_name,
        },
    )
    assert response.status_code == 201, response.text
    assert response.json()["status"] == "completed"


@pytest.fixture
def export_seed(
    security_seed: SecuritySeed,
    migration_database_url: str,
    tmp_path: Path,
) -> Iterator[SecuritySeed]:
    storage = LocalStorageProvider(tmp_path / "exports")
    app.dependency_overrides[get_export_storage] = lambda: storage
    confidential = (
        b"Date,Employee ID,Employee Name,Status,Department,Check In,Check Out,Total Hours\n"
        b'2026-08-01,EXP-001,"=HYPERLINK(""http://invalid.test"")",Present,'
        b"Engineering,09:00,17:00,8\n"
        b"2026-08-02,EXP-002,Fictional Two,Absent,Finance,,,0\n"
        b"2026-08-03,EXP-003,Fictional Three,Leave,Engineering,,,0\n"
    )
    public = (
        b"Date,Employee ID,Employee Name,Status,Department,Check In,Check Out,Total Hours\n"
        b"2026-08-04,PUBLIC-001,Public Fiction,WFH,Support,09:00,17:00,8\n"
    )
    low_entity = (
        b"Date,Employee ID,Employee Name,Status,Department\n"
        b"2026-08-05,LOW-001,Low Fiction,Present,Restricted Team\n"
    )
    with TestClient(app) as client:
        _upload(
            client,
            security_seed,
            confidential,
            logical_name="export-confidential",
            entity_id=security_seed.entity_high_id,
            classification=2,
        )
        _upload(
            client,
            security_seed,
            public,
            logical_name="export-public",
            entity_id=security_seed.entity_high_id,
            classification=0,
        )
        _upload(
            client,
            security_seed,
            low_entity,
            logical_name="export-low-entity",
            entity_id=security_seed.entity_low_id,
            classification=0,
        )
    try:
        yield security_seed
    finally:
        app.dependency_overrides.pop(get_export_storage, None)
        admin = create_engine(migration_database_url)
        with admin.begin() as connection:
            connection.execute(
                text("DELETE FROM documents WHERE product_id = :id"),
                {"id": security_seed.product_id},
            )
            connection.execute(text("TRUNCATE audit_events"))
        admin.dispose()


def _create_export(
    client: TestClient,
    seed: SecuritySeed,
    format_name: str,
    **overrides: Any,
) -> dict[str, Any]:
    payload = {
        "format": format_name,
        "dataset": "attendance",
        "entity_id": str(seed.entity_high_id),
        "module": "attendance",
        "classification": 2,
    }
    payload.update(overrides)
    response = client.post("/api/v1/exports", headers=_headers(seed), json=payload)
    assert response.status_code == 201, response.text
    assert response.json()["status"] == "completed", response.json()
    return response.json()


def _download(client: TestClient, seed: SecuritySeed, export_id: str) -> Any:
    response = client.get(f"/api/v1/exports/{export_id}/download", headers=_headers(seed))
    assert response.status_code == 200, response.text
    return response


@pytest.mark.integration
def test_json_xlsx_pdf_exports_are_consistent_and_preserve_lineage(
    export_seed: SecuritySeed,
) -> None:
    with TestClient(app) as client:
        jobs = {name: _create_export(client, export_seed, name) for name in ("json", "xlsx", "pdf")}
        status_response = client.get(
            f"/api/v1/exports/{jobs['json']['export_id']}", headers=_headers(export_seed)
        )
        downloads = {
            name: _download(client, export_seed, job["export_id"]) for name, job in jobs.items()
        }

    assert status_response.status_code == 200
    assert status_response.json()["record_count"] == 4

    json_payload = downloads["json"].json()
    json_ids = {row["employee_id"] for row in json_payload["records"]}
    assert json_ids == {"EXP-001", "EXP-002", "EXP-003", "PUBLIC-001"}
    assert "LOW-001" not in json_ids
    first = next(row for row in json_payload["records"] if row["employee_id"] == "EXP-001")
    assert first["source_file"] == "export-confidential.csv"
    assert first["source_row"] == 2
    assert first["source_record_id"]

    workbook = load_workbook(BytesIO(downloads["xlsx"].content), data_only=False)
    sheet = workbook["Attendance"]
    headers = [cell.value for cell in sheet[1]]
    xlsx_rows = [
        dict(zip(headers, row, strict=True)) for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    xlsx_ids = {row["employee_id"] for row in xlsx_rows}
    formula_name = next(
        row["employee_name"] for row in xlsx_rows if row["employee_id"] == "EXP-001"
    )
    assert formula_name.startswith("'=")
    assert sheet["C2"].data_type != "f"
    assert "Export Metadata" in workbook.sheetnames

    pdf_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(BytesIO(downloads["pdf"].content)).pages
    )
    pdf_ids = {employee_id for employee_id in json_ids if employee_id in pdf_text}
    assert json_ids == xlsx_ids == pdf_ids, pdf_text
    assert "export-confidential.csv" in pdf_text
    assert downloads["pdf"].content.startswith(b"%PDF")


@pytest.mark.integration
def test_export_filters_and_classification_do_not_leak_records(export_seed: SecuritySeed) -> None:
    with TestClient(app) as client:
        public_job = _create_export(
            client,
            export_seed,
            "json",
            classification=0,
        )
        filtered_job = _create_export(
            client,
            export_seed,
            "json",
            employee_id="EXP-002",
            status="absent",
            date_from="2026-08-02",
            date_to="2026-08-02",
        )
        public_rows = _download(client, export_seed, public_job["export_id"]).json()["records"]
        filtered_rows = _download(client, export_seed, filtered_job["export_id"]).json()["records"]
    assert [row["employee_id"] for row in public_rows] == ["PUBLIC-001"]
    assert [row["employee_id"] for row in filtered_rows] == ["EXP-002"]


@pytest.mark.integration
def test_export_scope_rbac_and_tenant_denials(
    export_seed: SecuritySeed,
    migration_database_url: str,
) -> None:
    base = {"format": "json", "dataset": "attendance", "module": "attendance"}
    with TestClient(app) as client:
        cross_tenant = client.post(
            "/api/v1/exports",
            headers=_headers(export_seed, tenant=export_seed.tenant_b_id),
            json={**base, "entity_id": str(export_seed.entity_high_id), "classification": 2},
        )
        entity_denied = client.post(
            "/api/v1/exports",
            headers=_headers(export_seed),
            json={**base, "entity_id": str(export_seed.entity_low_id), "classification": 2},
        )
        classification_denied = client.post(
            "/api/v1/exports",
            headers=_headers(export_seed),
            json={**base, "entity_id": str(export_seed.entity_high_id), "classification": 3},
        )
        module_denied = client.post(
            "/api/v1/exports",
            headers=_headers(export_seed),
            json={
                **base,
                "entity_id": str(export_seed.entity_high_id),
                "classification": 2,
                "module": "payroll",
            },
        )

    admin = create_engine(migration_database_url)
    with admin.begin() as connection:
        connection.execute(
            text(
                "DELETE FROM role_permissions rp USING permissions p, user_role_assignments ura "
                "WHERE rp.permission_id = p.id AND p.code = 'attendance:export' "
                "AND ura.role_id = rp.role_id AND ura.user_id = :user_id"
            ),
            {"user_id": export_seed.user_id},
        )
    admin.dispose()
    with TestClient(app) as client:
        rbac_denied = client.post(
            "/api/v1/exports",
            headers=_headers(export_seed),
            json={**base, "entity_id": str(export_seed.entity_high_id), "classification": 2},
        )

    assert cross_tenant.status_code == 403
    assert entity_denied.status_code == 403
    assert classification_denied.status_code == 403
    assert module_denied.status_code == 403
    assert rbac_denied.status_code == 403
    assert entity_denied.json() == {"detail": "Requested scope is unavailable"}


@pytest.mark.integration
def test_another_user_cannot_get_or_download_export(
    export_seed: SecuritySeed,
    migration_database_url: str,
) -> None:
    other_subject = f"export-other-{uuid4()}"
    admin = create_engine(migration_database_url)
    with Session(admin) as session:
        assignment = session.scalar(
            select(UserRoleAssignment).where(
                UserRoleAssignment.user_id == export_seed.user_id,
                UserRoleAssignment.entity_id == export_seed.entity_high_id,
            )
        )
        assert assignment is not None
        other = User(subject=other_subject, display_name="Other Export User")
        session.add(other)
        session.flush()
        session.add(
            UserRoleAssignment(
                user_id=other.id,
                role_id=assignment.role_id,
                product_id=assignment.product_id,
                tenant_id=assignment.tenant_id,
                entity_id=assignment.entity_id,
                module=assignment.module,
                classification_ceiling=assignment.classification_ceiling,
            )
        )
        session.commit()
        other_id = other.id

    try:
        with TestClient(app) as client:
            job = _create_export(client, export_seed, "json")
            other_headers = _headers(export_seed, subject=other_subject)
            status_response = client.get(
                f"/api/v1/exports/{job['export_id']}", headers=other_headers
            )
            download_response = client.get(
                f"/api/v1/exports/{job['export_id']}/download", headers=other_headers
            )
        assert status_response.status_code == 404
        assert download_response.status_code == 404
    finally:
        with Session(admin) as session:
            session.execute(text("DELETE FROM users WHERE id = :id"), {"id": other_id})
            session.commit()
        admin.dispose()


@pytest.mark.integration
def test_export_audit_and_expiration_cleanup(
    export_seed: SecuritySeed,
    migration_database_url: str,
) -> None:
    with TestClient(app) as client:
        job = _create_export(client, export_seed, "json")
        _download(client, export_seed, job["export_id"])

    admin = create_engine(migration_database_url)
    with admin.begin() as connection:
        audit_rows = connection.execute(
            text(
                "SELECT action, metadata FROM audit_events "
                "WHERE resource_id = :export_id ORDER BY created_at"
            ),
            {"export_id": job["export_id"]},
        ).all()
        storage_key = connection.scalar(
            text("SELECT storage_key FROM export_jobs WHERE id = :id"),
            {"id": job["export_id"]},
        )
        connection.execute(
            text("UPDATE export_jobs SET expires_at = now() - interval '1 second' WHERE id = :id"),
            {"id": job["export_id"]},
        )
    actions = [row.action for row in audit_rows]
    assert actions == ["export.requested", "export.completed", "export.downloaded"]
    assert all(row.metadata["export_id"] == job["export_id"] for row in audit_rows)
    assert all(row.metadata["request_id"] for row in audit_rows)
    assert [row.metadata["outcome"] for row in audit_rows] == [
        "requested",
        "completed",
        "downloaded",
    ]
    assert all("employee" not in json.dumps(row.metadata).lower() for row in audit_rows)

    storage = app.dependency_overrides[get_export_storage]()
    assert isinstance(storage, LocalStorageProvider)
    assert storage._resolve(storage_key).exists()
    with TestClient(app) as client:
        expired_download = client.get(
            f"/api/v1/exports/{job['export_id']}/download", headers=_headers(export_seed)
        )
        expired_status = client.get(
            f"/api/v1/exports/{job['export_id']}", headers=_headers(export_seed)
        )
    assert expired_download.status_code == 410
    assert expired_status.json()["status"] == "expired"
    assert not storage._resolve(storage_key).exists()
    admin.dispose()
